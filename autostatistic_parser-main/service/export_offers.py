
import os
import sys
from pprint import pprint
from datetime import datetime, timedelta, date

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Alignment, Font, Color, colors
from openpyxl.utils import get_column_letter
import smtplib
import mimetypes
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText

from logger import get_logger
from database import db_connect
from send_email import send_email

from common_config import LOGFILE_PATH, ROOT_PATH

logger = get_logger(__name__, LOGFILE_PATH)
connection = db_connect()

SHEET_TITLE = 'Список объявлений'
sheet_first_row = [
    'Дата попадания в базу',
    'Дата последней проверки',
    'Количество дней между датами',
    'Дубль',
    'Статус',
    'Коробка передач',
    'Заголовок',
    'Ссылка',
    'Цена (начальная)',
    'Цена (последняя)',
    'Год выпуска',
    'Пробег',
    'Кузов',
    'Цвет',
    'Марка',
    'Модель',
    'Объем двигателья',
    'Мощность',
    'Тип двигателя',
    'Привод',
    'Руль',
    'Состояние',
    'Владельцы',
    'ПТС',
    'Таможня',
    'Время владения',
    'Налог',
    'Гарантия',
    'Идентификатор',
    'Количество проверок',
    'Продавец',
    'Комплектация'
]
STATUS = {
    'None': 'Продается',
    'Sold': 'Продан',
    'Blocked': 'Заблокировано',
    'For_sale_again': 'Снова в продаже'
}

ALIGNMENT = Alignment(horizontal='left')
FONT = Font(bold=True)
FILL = PatternFill('solid', fgColor='FBE250')
FILL_TRACKING = PatternFill('solid', fgColor='FFB417')

NOW = datetime.now()
ADDR_FROM = 'root@367423-cc78716.tmweb.ru'
ADDR_TO = [
    'bulatovdm@yandex.ru',
    'm@vmonetu.ru'
]

if NOW.month < 10:
    MONTH = '0' + str(NOW.month)
else:
    MONTH = str(NOW.month)

if NOW.day < 10:
    DAY = '0' + str(NOW.day)
else:
    DAY = str(NOW.day)

DATE = DAY + '.' + MONTH + '.' + str(NOW.year)

dateFromExport = date.today() - timedelta(days=60)
dateFromExportTransformed = dateFromExport.strftime('%Y-%m-%d %H:%M:%S')

MESSAGE = 'Экспорт базы сервиса за ' + DATE
SUBJECT = 'Файл экспорта базы сервиса «Автостатистик» за ' + DATE
FILE_PATH = 'docs/autostatistic_export.xlsx'


def offers_export():
    """Exporting offers from a base.

    Returns:
        list: The list of offers.
    """

    sql = '''
        SELECT
          `l`.`id`,`l`.`createdon`,`l`.`title`,`l`.`url`,`l`.`price` AS price,`l`.`brand`,`l`.`model`,`l`.`release_year`,`l`.`run`,`l`.`body`,`l`.`color`,`l`.`engine_volume`,`l`.`power`,`l`.`engine_type`,`l`.`transmission`,`l`.`drive`,`l`.`hand_drive`,`l`.`condition`,`l`.`owners`,`l`.`pts`,`l`.`customs`,`l`.`possession_time`,`l`.`tax`,`l`.`guarantee`,`l`.`status` AS status, `l`.`checkedon`,`l`.`duplicated`,`l`.`inspections_number`,`l`.`seller_type`,`l`.`configuration`,`t`.`checkedon` AS checkedon_tracking,`t`.`price` AS price_tracking,`t`.`status` AS status_tracking
        FROM (
          (
            SELECT
              `l_sub`.`id`,`l_sub`.`createdon`,`l_sub`.`title`,`l_sub`.`url`,`l_sub`.`price`,`l_sub`.`brand`,`l_sub`.`model`,`l_sub`.`release_year`,`l_sub`.`run`,`l_sub`.`body`,`l_sub`.`color`,`l_sub`.`engine_volume`,`l_sub`.`power`,`l_sub`.`engine_type`,`l_sub`.`transmission`,`l_sub`.`drive`,`l_sub`.`hand_drive`,`l_sub`.`condition`,`l_sub`.`owners`,`l_sub`.`pts`,`l_sub`.`customs`,`l_sub`.`possession_time`,`l_sub`.`tax`,`l_sub`.`guarantee`,`l_sub`.`status`,`l_sub`.`checkedon`,`l_sub`.`duplicated`,`l_sub`.`inspections_number`, `l_sub`.`seller_type`,`l_sub`.`configuration`
            FROM
              `offers_list`AS `l_sub`
            WHERE
              `l_sub`.`checkedon` IS NOT NULL AND `l_sub`.`duplicated` IN(0,1) AND `l_sub`.`region` = "Санкт-Петербург"
            ORDER BY
              # `l_sub`.`createdon` ASC
              `l_sub`.`createdon` DESC
          ) AS l
        )
        LEFT JOIN
          `offers_tracking` AS `t`
        ON
          `l`.`id` = `t`.`item_id`
        WHERE
          `l`.`createdon` >= "''' + str(dateFromExportTransformed) + '''"
        ;
        '''
    cursor = connection.cursor()
    cursor.execute(sql)
    records = cursor.fetchall()
    data = []

    logger.info(f'Количество предложений для экспорта: {str(len(records))}.')

    for record in records:

        logger.info(f'Обрабатываем предложение с id: {str(record[0])}.')

        index = list_index_search(data, 'id', record[0])

        if record[26] == 0:
            duplicated = 'Нет'
        else:
            duplicated = 'Да'

        if record[24] is None:
            status = 'Продается'
        elif record[24] == 'Sold':
            status = 'Продан'
        elif record[24] == 'Blocked':
            status = 'Заблокировано'
        else:
            status = record[24]

        if record[4] == 0:
            price = ''
        else:
            price = record[4]
            price_last = price

        if index is False:
            item = {
                'createdon': record[1],
                'checkedon': record[25],
                'dates_interval': (record[25] - record[1]).days,
                'duplicated': duplicated,
                'status': status,
                'transmission': record[14],
                'title': record[2],
                'url': record[3],
                'price': price,
                'price_last': price_last,
                'release_year': record[7],
                'run': record[8],
                'body': record[9],
                'color': record[10],
                'brand': record[5],
                'model': record[6],
                'engine_volume': record[11],
                'power': record[12],
                'engine_type': record[13],
                'drive': record[15],
                'hand_drive': record[16],
                'condition': record[17],
                'owners': record[18],
                'pts': record[19],
                'customs': record[20],
                'possession_time': record[21],
                'tax': record[22],
                'guarantee': record[23],
                'id': record[0],
                'inspections_number': record[27],
                'seller_type': record[28],
                'configuration': record[29]
            }

            if record[30] is not None:

                if record[32] is None:
                    status = 'Продается'
                elif record[32] == 'Sold':
                    status = 'Продан'
                elif record[32] == 'Blocked':
                    status = 'Заблокировано'
                elif record[32] == 'For_sale_again':
                    status = 'Снова в продаже'
                else:
                    status = record[24]

                if record[31] == 0:
                    price = ''
                else:
                    price = record[31]

                    if item['price_last'] != price:
                        item['price_last'] = price

                item_tracking = {
                    'checkedon': record[30],
                    'price': price,
                    'status': status
                }
                item['tracking'] = [item_tracking]

            data.append(item)

        else:

            if record[32] is None:
                status = 'Продается'
            elif record[32] == 'Sold':
                status = 'Продан'
            elif record[32] == 'Blocked':
                status = 'Заблокировано'
            elif record[32] == 'For_sale_again':
                status = 'Снова в продаже'
            else:
                status = record[24]

            if record[31] == 0:
                price = ''
            else:
                price = record[31]

                if data[index]['price_last'] != price:
                    data[index]['price_last'] = price

            item_tracking = {
                'checkedon': record[30],
                'price': price,
                'status': status
            }

            if 'tracking' in data[index]:
                data[index]['tracking'].append(item_tracking)

    return data


def prepare_data(data):
    """Preparing the data for export.

    Args:
        data (list): The list of items for preparing.

    Returns:
        data: The list of rows and columns' width.
    """

    column_widths = []
    rows_to_export = []

    for i in range(len(data)):
        offer = data[i]
        row_to_export = []
        column_sheet = 0
        logger.info(
            f'Готовим данные для предложения с id: f{str(offer["id"])}.')

        for row in offer:
            column_sheet = column_sheet + 1
            price = offer['price']

            if row == 'tracking':

                for tracking_item in offer[row]:

                    checkedon_tracking = tracking_item['checkedon']
                    price_tracking = tracking_item['price']
                    status_tracking = tracking_item['status']

                    if price_tracking == price:
                        continue

                    if price_tracking != price:
                        price = price_tracking

                    try:
                        sheet_first_row[column_sheet - 1]
                    except:
                        sheet_first_row.append('Дата проверки')

                    column_sheet = column_sheet + 1
                    row_to_export.append(checkedon_tracking)
                    column_widths_set(
                        column_widths, column_sheet, checkedon_tracking)

                    try:
                        sheet_first_row[column_sheet - 1]
                    except:
                        sheet_first_row.append('Новая цена')

                    column_sheet = column_sheet + 1
                    row_to_export.append(price_tracking)
                    column_widths_set(
                        column_widths, column_sheet, price_tracking)

                    try:
                        sheet_first_row[column_sheet - 1]
                    except:
                        sheet_first_row.append('Новый статус')

                    column_sheet = column_sheet + 1
                    row_to_export.append(status_tracking)
                    column_widths_set(
                        column_widths, column_sheet, status_tracking)

            else:
                row_to_export.append(offer[row])
                column_widths_set(column_widths, column_sheet, offer[row])

        rows_to_export.append(row_to_export)

    return [rows_to_export, column_widths]


def xlsx_export(data):
    """Exporting the data to the file

    Args:
        data (list): The data list.
    """

    rows = data[0]
    column_widths = data[1]
    wb = openpyxl.Workbook(write_only=True)
    sheet = wb.create_sheet()
    sheet.title = SHEET_TITLE
    header_to_export = []

    for i in range(len(sheet_first_row)):
        cell = WriteOnlyCell(sheet)
        cell.font = FONT

        if i > 31:
            cell.fill = FILL_TRACKING
        else:
            cell.fill = FILL

        cell.alignment = ALIGNMENT
        cell.value = sheet_first_row[i]
        column_widths_set(column_widths, i + 1, sheet_first_row[i])
        header_to_export.append(cell)

    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = 'A1:' + \
        str(get_column_letter(len(sheet_first_row))) + str(len(rows))

    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(
            i + 1)].width = column_width + 6

    sheet.append(header_to_export)

    for row in rows:
        sheet.append(row)

    wb.save('docs/autostatistic_export.xlsx')

    return


def column_widths_set(column_widths, column_sheet, value):
    """Setting the column' width.

    Args:
        column_widths (list): The list of columns' widths.
        column_sheet (int): The sheet's index.
        value (mixed): The cell value.
    """

    try:
        if len(str(value)) > column_widths[column_sheet - 1]:
            column_widths[column_sheet - 1] = len(str(value))

    except IndexError:
        column_widths.append(len(str(value)))


def list_index_search(lst, key, value):
    """Searching the index of the list by key and value.

    Args:
        lst (list): The input array.
        key (str): The specified key.
        value (mixed): The value to search.

    Returns:
        int: The list index.
    """

    for i in range(len(lst)):
        if lst[i][key] == value:
            return i
    return False


def remove_file(file):
    """Removing the file.

    Args:
        file (str): The file to remove.
    """

    path = os.path.abspath(file)
    os.remove(path)


data = offers_export()
data = prepare_data(data)
xlsx_export(data)
send_email(SUBJECT, ADDR_FROM, ADDR_TO, MESSAGE, FILE_PATH)
remove_file(FILE_PATH)
